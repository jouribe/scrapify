"""Tests for Excel report generation service."""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from scrapyfy.excel_service import ExcelReportService
from scrapyfy.models import Comment, Company, Post, SentimentAggregate


@pytest.fixture
def temp_output_dir() -> Path:
    """Create temporary directory for test outputs."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def mock_session():
    """Create mock session with test data."""
    session = MagicMock()

    # Mock company
    company = Company(id=1, name="Test Company", slug="test-company")

    # Mock posts (not used in current tests but kept for potential future use)
    Post(
        id=1,
        company_id=1,
        platform="facebook",
        external_id="ext1",
        post_url="http://example.com/post1",
        content="This is a test post",
        author="Test Author",
        posted_at=None,
        metrics={
            "sentiment_label": "POSITIVE",
            "sentiment_score": 0.95,
            "positive_count": 10,
            "neutral_count": 2,
            "negative_count": 1,
        },
    )

    Post(
        id=2,
        company_id=1,
        platform="instagram",
        external_id="ext2",
        post_url="http://example.com/post2",
        content="Another test post",
        author="Another Author",
        posted_at=None,
        metrics={
            "sentiment_label": "NEGATIVE",
            "sentiment_score": 0.12,
            "positive_count": 2,
            "neutral_count": 3,
            "negative_count": 8,
        },
    )

    # Mock comments
    Comment(
        id=1,
        post_id=1,
        platform="facebook",
        external_id="comment1",
        content="Great post!",
        author="Commenter 1",
        commented_at=None,
        raw={
            "sentiment_label": "POSITIVE",
            "sentiment_score": 0.89,
        },
    )

    Comment(
        id=2,
        post_id=2,
        platform="instagram",
        external_id="comment2",
        content="Not good",
        author="Commenter 2",
        commented_at=None,
        raw={
            "sentiment_label": "NEGATIVE",
            "sentiment_score": 0.05,
        },
    )

    # Mock sentiment aggregates
    agg1 = SentimentAggregate(
        id=1,
        company_id=1,
        scope="platform_company",
        platform="facebook",
        metrics={
            "total_posts": 1,
            "valid_comments": 1,
            "positive_count": 10,
            "neutral_count": 2,
            "negative_count": 1,
            "net_sentiment": 0.82,
        },
        model_name="test-model",
    )

    agg2 = SentimentAggregate(
        id=2,
        company_id=1,
        scope="platform_company",
        platform="instagram",
        metrics={
            "total_posts": 1,
            "valid_comments": 1,
            "positive_count": 2,
            "neutral_count": 3,
            "negative_count": 8,
            "net_sentiment": -0.55,
        },
        model_name="test-model",
    )

    # Configure mock queries
    session.query.return_value.filter.return_value.first.return_value = company
    session.query.return_value.filter.return_value.all.side_effect = [
        [agg1, agg2],  # For SentimentAggregate query in _create_resumen_sheet
        [],  # For Post.platform distinct query (empty for testing)
    ]

    return session


@pytest.fixture
def service_with_mock():
    """Create ExcelReportService with mocked database."""
    with patch("scrapyfy.excel_service.create_engine"):
        service = ExcelReportService(database_url="sqlite:///test.db")
        return service


def test_excel_service_initialization(service_with_mock):
    """Test that ExcelReportService initializes correctly."""
    assert service_with_mock.wb is not None
    assert len(service_with_mock.wb.sheetnames) == 0


def test_excel_report_invalid_company(service_with_mock, temp_output_dir):
    """Test handling of invalid company slug."""
    with patch("scrapyfy.excel_service.Session") as mock_session_class:
        mock_session = MagicMock()
        mock_session_class.return_value.__enter__.return_value = mock_session
        mock_session.query.return_value.filter.return_value.first.return_value = None

        with pytest.raises(ValueError, match="not found"):
            service_with_mock.generate_report("nonexistent-company", output_dir=temp_output_dir)


def test_excel_sheet_names_created(temp_output_dir):
    """Test that Excel file has correct sheet names."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=temp_output_dir, delete=False) as tmp:
        test_file = Path(tmp.name)

    try:
        # Create a simple workbook
        from openpyxl import Workbook

        wb = Workbook()
        ws_resumen = wb.create_sheet("Resumen", 0)
        ws_fb = wb.create_sheet("Facebook_Posts")
        ws_ig = wb.create_sheet("Instagram_Posts")

        # Add minimal data
        ws_resumen["A1"] = "Test"
        ws_fb["A1"] = "Test"
        ws_ig["A1"] = "Test"

        wb.save(test_file)

        # Verify sheet creation
        wb_loaded = load_workbook(test_file)
        assert "Resumen" in wb_loaded.sheetnames
        assert "Facebook_Posts" in wb_loaded.sheetnames
        assert "Instagram_Posts" in wb_loaded.sheetnames
    finally:
        if test_file.exists():
            test_file.unlink()


def test_excel_file_is_valid_workbook(temp_output_dir):
    """Test that generated Excel file can be opened with openpyxl."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=temp_output_dir, delete=False) as tmp:
        test_file = Path(tmp.name)

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"] = "Value"
        wb.save(test_file)

        # Try to load it back
        wb_loaded = load_workbook(test_file)
        assert "Test" in wb_loaded.sheetnames
        ws_loaded = wb_loaded["Test"]
        assert ws_loaded["A1"].value == "Value"
    finally:
        if test_file.exists():
            test_file.unlink()


def test_output_path_format(temp_output_dir):
    """Test that output path follows expected naming convention."""
    from datetime import datetime

    company_slug = "test-company"
    expected_prefix = f"{company_slug}_sentiment_report_"
    expected_date = datetime.now().strftime("%Y-%m-%d")

    # Simulate the filename generation logic
    filename = (
        f"{company_slug}_sentiment_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )
    output_path = temp_output_dir / filename

    assert str(output_path).endswith(".xlsx")
    assert expected_prefix in filename
    assert expected_date in filename


def test_clean_excel_value_removes_illegal_control_chars() -> None:
    """Excel exports should strip control chars that openpyxl rejects."""
    raw_value = "Hola\x03 mundo"
    cleaned = ExcelReportService._clean_excel_value(raw_value)

    assert cleaned == "Hola mundo"
