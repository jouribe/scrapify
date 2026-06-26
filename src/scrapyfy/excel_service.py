"""Excel report generation service for sentiment analysis and social media data."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from dateutil.parser import parse as parse_datetime
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from scrapyfy.logging_config import get_logger
from scrapyfy.models import Comment, Company, Post, SentimentAggregate

logger = get_logger(__name__)

# Color scheme
COLOR_HEADER = "4472C4"
COLOR_POSITIVE = "70AD47"
COLOR_NEGATIVE = "FF5050"
COLOR_NEUTRAL = "A9A9A9"
COLOR_NO_DATA = "D9E2F3"
COLOR_BORDER = "D3D3D3"
COLOR_ALT_ROW = "F0F0F0"
COLOR_TEXT_DARK = "203864"


class ExcelReportService:
    """Generates professional Excel reports for social media sentiment analysis."""

    def __init__(self, database_url: str):
        """Initialize the service with database connection."""
        self.engine = create_engine(database_url, future=True, pool_pre_ping=True)
        self._reset_workbook()

    def _reset_workbook(self) -> None:
        """Create a fresh workbook for each export operation."""
        self.wb = Workbook()
        self._remove_default_sheet()

    def _remove_default_sheet(self) -> None:
        """Remove the default empty sheet created by openpyxl."""
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

    @staticmethod
    def _clean_excel_value(value: Any) -> Any:
        """Strip control characters that Excel/openpyxl cannot write to a worksheet."""
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    @staticmethod
    def _coerce_datetime(value: Any) -> datetime | None:
        """Convert multiple raw date formats to a Python datetime for Excel export."""
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value

        try:
            if isinstance(value, (int, float)):
                timestamp = float(value)
                if timestamp > 1_000_000_000_000:
                    timestamp /= 1000
                return datetime.fromtimestamp(timestamp, tz=UTC).replace(tzinfo=None)

            text_value = str(value).strip()
            if text_value.isdigit():
                timestamp = float(text_value)
                if timestamp > 1_000_000_000_000:
                    timestamp /= 1000
                return datetime.fromtimestamp(timestamp, tz=UTC).replace(tzinfo=None)

            parsed = parse_datetime(text_value)
            return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
        except (TypeError, ValueError, OverflowError):
            return None

    @classmethod
    def _post_display_datetime(cls, post: Post) -> datetime | str:
        """Resolve the best available datetime for display, including raw fallbacks."""
        if post.posted_at:
            return post.posted_at.replace(tzinfo=None) if post.posted_at.tzinfo else post.posted_at

        raw = dict(post.raw or {})
        for key in (
            "createTimeISO",
            "createdAt",
            "takenAt",
            "time",
            "publishedAt",
            "postedAtISO",
            "postedAt",
            "date",
            "timestamp",
            "postedAtTimestamp",
        ):
            candidate = cls._coerce_datetime(raw.get(key))
            if candidate is not None:
                return candidate

        return ""

    def generate_report(
        self,
        company_slug: str,
        output_dir: Path | str = ".",
    ) -> Path:
        """
        Generate complete Excel report for a company.

        Args:
            company_slug: Slug identifier for the company (e.g., "caja-arequipa")
            output_dir: Directory to save the Excel file

        Returns:
            Path to generated Excel file
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        self._reset_workbook()

        with Session(self.engine) as session:
            company = self._get_company(session, company_slug)
            if not company:
                raise ValueError(f"Company with slug '{company_slug}' not found")

            self._create_resumen_sheet(session, company)
            self._create_platform_sheets(session, company)

        filename = (
            f"{company_slug}_sentiment_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        output_path = output_dir / filename
        self.wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")

        return output_path

    def _get_company(self, session: Session, slug: str) -> Company | None:
        """Retrieve company by slug."""
        return session.query(Company).filter(Company.slug == slug).first()

    def _create_resumen_sheet(self, session: Session, company: Company) -> None:
        """Create the Resumen (summary) sheet with charts and overview."""
        ws = self.wb.create_sheet("Resumen", 0)

        # Title
        ws["A1"] = self._clean_excel_value(f"Reporte de Sentimiento - {company.name}")
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
        )
        ws.merge_cells("A1:H1")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Metadata
        row = 3
        ws[f"A{row}"] = "Empresa:"
        ws[f"B{row}"] = self._clean_excel_value(company.name)
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font()

        row += 1
        ws[f"A{row}"] = "Generado:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font()

        # Summary table header
        row = 6
        summary_headers = ["Red Social", "Posts", "Comentarios", "Pos", "Neu", "Neg", "Sentimiento"]
        for col_idx, header_text in enumerate(summary_headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header_text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch global aggregates per platform
        aggregates = (
            session.query(SentimentAggregate)
            .filter(
                SentimentAggregate.company_id == company.id,
                SentimentAggregate.scope == "platform_company",
            )
            .all()
        )

        row = 7
        for agg in sorted(aggregates, key=lambda a: a.platform or ""):
            metrics = agg.metrics or {}
            pos_count = metrics.get("positive_count", 0)
            neu_count = metrics.get("neutral_count", 0)
            neg_count = metrics.get("negative_count", 0)
            net_sentiment = metrics.get("net_sentiment", 0.0)

            ws.cell(row=row, column=1).value = self._clean_excel_value(agg.platform)
            ws.cell(row=row, column=2).value = metrics.get("total_posts", 0)
            ws.cell(row=row, column=3).value = metrics.get("valid_comments", 0)
            ws.cell(row=row, column=4).value = pos_count
            ws.cell(row=row, column=5).value = neu_count
            ws.cell(row=row, column=6).value = neg_count
            ws.cell(row=row, column=7).value = f"{net_sentiment:.2f}"

            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=COLOR_ALT_ROW,
                        end_color=COLOR_ALT_ROW,
                        fill_type="solid",
                    )
            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 18
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 14

        # Create charts
        chart_row = 7 + len(aggregates) + 2
        self._add_sentiment_charts(ws, aggregates, chart_row)

    def _add_sentiment_charts(
        self, ws: Any, aggregates: list[SentimentAggregate], start_row: int
    ) -> None:
        """Add sentiment analysis charts to the summary sheet."""
        if not aggregates:
            return

        # Bar chart: Positive/Negative/Neutral per platform
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "Sentimiento por Red Social"
        bar_chart.gapWidth = 70
        bar_chart.y_axis.title = "Cantidad de Comentarios"
        bar_chart.x_axis.title = "Red Social"
        bar_chart.legend.position = "r"
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = False
        data_end_row = 6 + len(aggregates)

        # Add data series
        pos_values = Reference(ws, min_col=4, min_row=6, max_row=data_end_row)
        neu_values = Reference(ws, min_col=5, min_row=6, max_row=data_end_row)
        neg_values = Reference(ws, min_col=6, min_row=6, max_row=data_end_row)
        categories = Reference(ws, min_col=1, min_row=7, max_row=data_end_row)

        bar_chart.add_data(pos_values, titles_from_data=True)
        bar_chart.add_data(neu_values, titles_from_data=True)
        bar_chart.add_data(neg_values, titles_from_data=True)
        bar_chart.set_categories(categories)

        # Color the bars
        series_pos = bar_chart.series[0]
        series_pos.graphicalProperties.solidFill = COLOR_POSITIVE
        series_neu = bar_chart.series[1]
        series_neu.graphicalProperties.solidFill = COLOR_NEUTRAL
        series_neg = bar_chart.series[2]
        series_neg.graphicalProperties.solidFill = COLOR_NEGATIVE

        # Position chart
        ws.add_chart(bar_chart, f"A{start_row}")

        # Pie chart: Distribution by platform
        pie_chart = PieChart()
        pie_chart.title = "Distribución de Comentarios por Red"
        pie_chart.legend.position = "r"

        comment_counts = Reference(ws, min_col=3, min_row=6, max_row=data_end_row)
        pie_categories = Reference(ws, min_col=1, min_row=7, max_row=data_end_row)

        pie_chart.add_data(comment_counts, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showLegendKey = False
        pie_chart.dataLabels.showCatName = True
        pie_chart.dataLabels.separator = " - "

        if pie_chart.series:
            pie_series = pie_chart.series[0]
            solid_colors = ["4472C4", "ED7D31", "70AD47", "A5A5A5", "FFC000"]
            pie_series.data_points = []
            for idx, _ in enumerate(aggregates):
                point = DataPoint(idx=idx)
                point.graphicalProperties.solidFill = solid_colors[idx % len(solid_colors)]
                pie_series.data_points.append(point)

        ws.add_chart(pie_chart, f"E{start_row}")

    def _create_platform_sheets(self, session: Session, company: Company) -> None:
        """Create sheets for each platform with posts and comments data."""
        platforms = (
            session.query(Post.platform)
            .filter(Post.company_id == company.id)
            .distinct(Post.platform)
            .order_by(Post.platform)
            .all()
        )

        for (platform,) in platforms:
            self._create_platform_sheet(session, company, platform)

    def _create_platform_sheet(self, session: Session, company: Company, platform: str) -> None:
        """Create a sheet for posts and comments of a specific platform."""
        sheet_name = f"{platform.capitalize()}_Posts"
        # Ensure sheet name is valid (max 31 chars)
        sheet_name = sheet_name[:31]

        ws = self.wb.create_sheet(sheet_name)

        # Posts section header
        ws["A1"] = f"Posts - {platform.capitalize()}"
        ws["A1"].font = Font(size=12, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
        )
        ws.merge_cells("A1:H1")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22

        # Posts table headers
        post_headers = [
            "Autor",
            "Contenido",
            "URL",
            "Fecha",
            "Sentimiento",
            "Confianza",
            "Positivos",
            "Neutrales",
            "Negativos",
        ]

        for col_idx, header_text in enumerate(post_headers, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header_text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch posts
        posts = (
            session.query(Post)
            .filter(Post.company_id == company.id, Post.platform == platform)
            .order_by(Post.posted_at.desc())
            .all()
        )

        row = 3
        for post in posts:
            metrics = post.metrics or {}
            sentiment = metrics.get("sentiment_label", "N/A")
            confidence = metrics.get("sentiment_score", 0.0)
            pos_count = metrics.get("positive_count", 0)
            neu_count = metrics.get("neutral_count", 0)
            neg_count = metrics.get("negative_count", 0)

            ws.cell(row=row, column=1).value = self._clean_excel_value(post.author or "")
            ws.cell(row=row, column=2).value = self._clean_excel_value(
                (post.content or "")[:100]
            )
            ws.cell(row=row, column=3).value = self._clean_excel_value(post.post_url or "")
            post_display_datetime = self._post_display_datetime(post)
            date_cell = ws.cell(row=row, column=4)
            date_cell.value = post_display_datetime
            if isinstance(post_display_datetime, datetime):
                date_cell.number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row=row, column=5).value = sentiment
            ws.cell(row=row, column=6).value = f"{confidence:.2f}"
            ws.cell(row=row, column=7).value = pos_count
            ws.cell(row=row, column=8).value = neu_count
            ws.cell(row=row, column=9).value = neg_count

            # Color sentiment column
            sentiment_cell = ws.cell(row=row, column=5)
            if sentiment == "POSITIVE":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_POSITIVE,
                    end_color=COLOR_POSITIVE,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)
            elif sentiment == "NEGATIVE":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_NEGATIVE,
                    end_color=COLOR_NEGATIVE,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)
            elif sentiment == "NEUTRAL":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_NEUTRAL,
                    end_color=COLOR_NEUTRAL,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)

            # Alternate row coloring for other columns
            if row % 2 == 0:
                for col in range(1, 10):
                    if col != 5:  # Skip sentiment column (already colored)
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=COLOR_ALT_ROW,
                            end_color=COLOR_ALT_ROW,
                            fill_type="solid",
                        )

            row += 1

        # Set column widths for posts
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12
        for col in ["G", "H", "I"]:
            ws.column_dimensions[col].width = 12

        # Freeze panes
        ws.freeze_panes = "A3"

        # Comments section
        comments_start_row = row + 2
        ws[f"A{comments_start_row}"] = f"Comentarios - {platform.capitalize()}"
        ws[f"A{comments_start_row}"].font = Font(size=11, bold=True, color="FFFFFF")
        ws[f"A{comments_start_row}"].fill = PatternFill(
            start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
        )
        ws.merge_cells(f"A{comments_start_row}:E{comments_start_row}")
        ws.row_dimensions[comments_start_row].height = 18

        # Comments table headers
        comment_headers = ["Autor", "Comentario", "Sentimiento", "Confianza", "Fecha"]
        for col_idx, header_text in enumerate(comment_headers, start=1):
            cell = ws.cell(row=comments_start_row + 1, column=col_idx)
            cell.value = header_text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch comments
        comments = (
            session.query(Comment)
            .join(Post)
            .filter(Post.company_id == company.id, Post.platform == platform)
            .order_by(Comment.commented_at.desc())
            .all()
        )

        row = comments_start_row + 2
        for comment in comments:
            raw = comment.raw or {}
            sentiment = raw.get("sentiment_label", "N/A")
            confidence = raw.get("sentiment_score", 0.0)

            ws.cell(row=row, column=1).value = self._clean_excel_value(comment.author or "")
            # Truncate for visibility
            ws.cell(row=row, column=2).value = self._clean_excel_value((comment.content or "")[:80])
            ws.cell(row=row, column=3).value = sentiment
            ws.cell(row=row, column=4).value = f"{confidence:.2f}"
            ws.cell(row=row, column=5).value = (
                comment.commented_at.strftime("%Y-%m-%d %H:%M") if comment.commented_at else ""
            )

            # Color sentiment column
            sentiment_cell = ws.cell(row=row, column=3)
            if sentiment == "POSITIVE":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_POSITIVE,
                    end_color=COLOR_POSITIVE,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)
            elif sentiment == "NEGATIVE":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_NEGATIVE,
                    end_color=COLOR_NEGATIVE,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)
            elif sentiment == "NEUTRAL":
                sentiment_cell.fill = PatternFill(
                    start_color=COLOR_NEUTRAL,
                    end_color=COLOR_NEUTRAL,
                    fill_type="solid",
                )
                sentiment_cell.font = Font(color="FFFFFF", bold=True)

            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, 6):
                    if col != 3:  # Skip sentiment column
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=COLOR_ALT_ROW,
                            end_color=COLOR_ALT_ROW,
                            fill_type="solid",
                        )

            row += 1

        # Set column widths for comments
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18

    def generate_platform_report(
        self,
        platform: str,
        company_slugs: list[str],
        output_dir: Path | str = ".",
    ) -> Path:
        """Generate a platform-specific workbook for multiple companies."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        self._reset_workbook()

        normalized_platform = platform.strip().lower()
        with Session(self.engine) as session:
            companies = self._get_companies(session, company_slugs)
            posts = (
                session.query(Post)
                .join(Company, Post.company_id == Company.id)
                .filter(
                    Post.platform == normalized_platform,
                    Company.slug.in_([company.slug for company in companies]),
                )
                .order_by(Company.name, Post.posted_at.desc())
                .all()
            )

            if not posts:
                raise ValueError(
                    "No posts found for the selected companies and platform: "
                    f"platform={normalized_platform}"
                )

            self._create_multi_company_summary_sheet(normalized_platform, companies, posts)
            self._create_posts_export_sheet(
                title=f"{normalized_platform.capitalize()} - Posts",
                sheet_name="Posts",
                posts=posts,
            )

            for company in companies:
                company_posts = [post for post in posts if post.company_id == company.id]
                if company_posts:
                    self._create_posts_export_sheet(
                        title=f"{company.name} - {normalized_platform.capitalize()}",
                        sheet_name=self._normalize_sheet_name(company.name),
                        posts=company_posts,
                    )

        filename = (
            f"{normalized_platform}_multi_company_report_"
            f"{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )
        output_path = output_dir / filename
        self.wb.save(output_path)
        logger.info("Platform Excel report saved: %s", output_path)
        return output_path

    def _get_companies(self, session: Session, company_slugs: list[str]) -> list[Company]:
        """Resolve companies by slug while preserving the requested order."""
        normalized_slugs = [slug.strip() for slug in company_slugs if slug.strip()]
        if not normalized_slugs:
            raise ValueError("At least one company slug is required")

        companies = (
            session.query(Company)
            .filter(Company.slug.in_(normalized_slugs))
            .order_by(Company.name)
            .all()
        )
        company_by_slug = {company.slug: company for company in companies}

        missing = [slug for slug in normalized_slugs if slug not in company_by_slug]
        if missing:
            raise ValueError(f"Companies not found: {', '.join(missing)}")

        return [company_by_slug[slug] for slug in normalized_slugs]

    def _create_multi_company_summary_sheet(
        self,
        platform: str,
        companies: list[Company],
        posts: list[Post],
    ) -> None:
        """Create a benchmark-style summary sheet for one platform across companies."""
        ws = self.wb.create_sheet("Resumen", 0)
        ws.sheet_view.showGridLines = False

        title = f"Benchmark {platform.capitalize()} - Resumen por marca"
        ws["A1"] = self._clean_excel_value(title)
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color=COLOR_HEADER,
            end_color=COLOR_HEADER,
            fill_type="solid",
        )
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:Q1")
        ws.row_dimensions[1].height = 28

        ws["A2"] = "Empresas"
        ws["B2"] = self._clean_excel_value(", ".join(company.name for company in companies))
        ws["A3"] = "Generado"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A2"].font = ws["A3"].font = Font(bold=True)

        headers = [
            "company",
            "platform",
            "posts",
            "total_likes",
            "total_comments",
            "total_shares",
            "total_views",
            "total_saves",
            "avg_likes",
            "avg_comments",
            "avg_shares",
            "avg_views",
            "avg_saves",
            "positive_posts",
            "neutral_posts",
            "negative_posts",
            "no_data_posts",
        ]
        header_row = 5
        for column_index, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=column_index, value=header)

        row = header_row + 1
        for company in companies:
            company_posts = [post for post in posts if post.company_id == company.id]
            likes = [self._metric_value(post, "likes") for post in company_posts]
            comments = [self._metric_value(post, "comments") for post in company_posts]
            shares = [self._metric_value(post, "shares") for post in company_posts]
            views = [self._metric_value(post, "views") for post in company_posts]
            saves = [
                self._metric_value(
                    post,
                    "saves",
                    ("saves", "saveCount", "savesCount"),
                )
                for post in company_posts
            ]

            sentiment_labels = [
                str((post.metrics or {}).get("sentiment_label", "NO_DATA")).upper()
                for post in company_posts
            ]

            values = [
                company.name,
                platform,
                len(company_posts),
                sum(likes),
                sum(comments),
                sum(shares),
                sum(views),
                sum(saves),
                round(sum(likes) / len(company_posts), 2) if company_posts else 0,
                round(sum(comments) / len(company_posts), 2) if company_posts else 0,
                round(sum(shares) / len(company_posts), 2) if company_posts else 0,
                round(sum(views) / len(company_posts), 2) if company_posts else 0,
                round(sum(saves) / len(company_posts), 2) if company_posts else 0,
                sum(label == "POSITIVE" for label in sentiment_labels),
                sum(label == "NEUTRAL" for label in sentiment_labels),
                sum(label == "NEGATIVE" for label in sentiment_labels),
                sum(label == "NO_DATA" for label in sentiment_labels),
            ]

            for column_index, value in enumerate(values, start=1):
                ws.cell(
                    row=row,
                    column=column_index,
                    value=self._clean_excel_value(value),
                )
            row += 1

        self._apply_table_styles(
            ws,
            header_row=header_row,
            last_row=row - 1,
            column_widths={
                1: 20,
                2: 12,
                3: 10,
                4: 12,
                5: 14,
                6: 12,
                7: 12,
                8: 12,
                9: 12,
                10: 14,
                11: 12,
                12: 12,
                13: 12,
                14: 13,
                15: 13,
                16: 13,
                17: 13,
            },
        )

    def _create_posts_export_sheet(
        self,
        title: str,
        sheet_name: str,
        posts: list[Post],
    ) -> None:
        """Create a polished tabular export sheet similar to the sample workbook."""
        ws = self.wb.create_sheet(sheet_name[:31])
        ws.sheet_view.showGridLines = False

        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color=COLOR_HEADER,
            end_color=COLOR_HEADER,
            fill_type="solid",
        )
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:R1")
        ws.row_dimensions[1].height = 24

        headers = [
            "company",
            "date",
            "content",
            "likes",
            "comments",
            "shares",
            "views",
            "saves",
            "platform",
            "content_type",
            "url",
            "source_account",
            "sentiment",
            "confidence",
            "positive_comments",
            "neutral_comments",
            "negative_comments",
            "valid_comments",
        ]
        header_row = 3
        for column_index, header in enumerate(headers, start=1):
            ws.cell(row=header_row, column=column_index, value=header)

        row = header_row + 1
        for post in posts:
            metrics = dict(post.metrics or {})
            raw = dict(post.raw or {})
            sentiment = str(metrics.get("sentiment_label", "NO_DATA")).upper()
            post_display_datetime = self._post_display_datetime(post)
            values = [
                post.company.name if post.company else "",
                post_display_datetime,
                post.content or "",
                self._metric_value(post, "likes"),
                self._metric_value(post, "comments"),
                self._metric_value(post, "shares"),
                self._metric_value(post, "views"),
                self._metric_value(post, "saves", ("saves", "saveCount", "savesCount")),
                post.platform,
                self._detect_content_type(raw),
                post.post_url or "",
                post.author or "",
                sentiment,
                round(self._safe_number(metrics.get("sentiment_score")), 3),
                int(self._safe_number(metrics.get("positive_count"))),
                int(self._safe_number(metrics.get("neutral_count"))),
                int(self._safe_number(metrics.get("negative_count"))),
                int(self._safe_number(metrics.get("valid_comments"))),
            ]

            for column_index, value in enumerate(values, start=1):
                safe_value = self._clean_excel_value(value)
                cell = ws.cell(row=row, column=column_index, value=safe_value)
                if column_index == 2 and isinstance(safe_value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"
                if column_index == 11 and safe_value:
                    cell.hyperlink = str(safe_value)
                    cell.style = "Hyperlink"

            self._style_sentiment_cell(ws.cell(row=row, column=13), sentiment)
            row += 1

        self._apply_table_styles(
            ws,
            header_row=header_row,
            last_row=row - 1,
            column_widths={
                1: 18,
                2: 18,
                3: 60,
                4: 10,
                5: 10,
                6: 10,
                7: 10,
                8: 10,
                9: 12,
                10: 14,
                11: 40,
                12: 20,
                13: 12,
                14: 11,
                15: 14,
                16: 14,
                17: 14,
                18: 13,
            },
            wrapped_columns={3, 11},
        )

    @staticmethod
    def _metric_value(post: Post, key: str, raw_keys: tuple[str, ...] | None = None) -> float:
        """Extract a numeric metric from post.metrics or post.raw."""
        raw_keys = raw_keys or ()
        metrics = dict(post.metrics or {})
        raw = dict(post.raw or {})

        if key in metrics:
            return ExcelReportService._safe_number(metrics.get(key))

        for raw_key in raw_keys:
            if raw_key in raw:
                return ExcelReportService._safe_number(raw.get(raw_key))

        return 0.0

    @staticmethod
    def _safe_number(value: Any) -> float:
        """Convert metric values to float safely."""
        if value in (None, "", False):
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _detect_content_type(raw: dict[str, Any]) -> str:
        """Infer a content type label from the raw payload."""
        for key in ("type", "contentType", "postType", "mediaType", "__typename"):
            value = raw.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
        return "post"

    @staticmethod
    def _normalize_sheet_name(name: str) -> str:
        """Make a workbook-safe sheet name."""
        invalid_chars = {"\\", "/", "*", "[", "]", ":", "?"}
        sanitized = "".join("-" if char in invalid_chars else char for char in name)
        return sanitized[:31]

    @staticmethod
    def _style_sentiment_cell(cell: Any, sentiment: str) -> None:
        """Apply a color scheme to the sentiment column."""
        if sentiment == "POSITIVE":
            cell.fill = PatternFill(
                start_color=COLOR_POSITIVE,
                end_color=COLOR_POSITIVE,
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif sentiment == "NEGATIVE":
            cell.fill = PatternFill(
                start_color=COLOR_NEGATIVE,
                end_color=COLOR_NEGATIVE,
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif sentiment == "NEUTRAL":
            cell.fill = PatternFill(
                start_color=COLOR_NEUTRAL,
                end_color=COLOR_NEUTRAL,
                fill_type="solid",
            )
            cell.font = Font(color="FFFFFF", bold=True)
        elif sentiment == "NO_DATA":
            cell.fill = PatternFill(
                start_color=COLOR_NO_DATA,
                end_color=COLOR_NO_DATA,
                fill_type="solid",
            )
            cell.font = Font(color=COLOR_TEXT_DARK, bold=True)

    @staticmethod
    def _apply_table_styles(
        ws: Any,
        header_row: int,
        last_row: int,
        column_widths: dict[int, int],
        wrapped_columns: set[int] | None = None,
    ) -> None:
        """Apply consistent styling to export tables."""
        wrapped_columns = wrapped_columns or set()
        border = Border(
            left=Side(style="thin", color=COLOR_BORDER),
            right=Side(style="thin", color=COLOR_BORDER),
            top=Side(style="thin", color=COLOR_BORDER),
            bottom=Side(style="thin", color=COLOR_BORDER),
        )

        max_column = max(column_widths)
        ws.freeze_panes = f"A{header_row + 1}"
        last_col_letter = get_column_letter(max_column)
        last_data_row = max(last_row, header_row)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row}"

        for column_index, width in column_widths.items():
            ws.column_dimensions[get_column_letter(column_index)].width = width

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=COLOR_HEADER,
                end_color=COLOR_HEADER,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_index in range(header_row + 1, last_row + 1):
            for column_index in range(1, max_column + 1):
                cell = ws.cell(row=row_index, column=column_index)
                if cell.fill.fill_type is None and row_index % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=COLOR_ALT_ROW,
                        end_color=COLOR_ALT_ROW,
                        fill_type="solid",
                    )
                cell.border = border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=column_index in wrapped_columns,
                )
